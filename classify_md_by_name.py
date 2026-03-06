#!/usr/bin/env python3
"""Classify Word and Excel files into name-based folders using content-based detection.

By default, files are copied into folders named after detected names.
Use --move to move files instead.

Examples:
    python classify_md_by_name.py . --dry-run
    python classify_md_by_name.py . --move
"""

from __future__ import annotations

import argparse
import re
import shutil
import sys
from pathlib import Path

try:
    from docx import Document
except ImportError:
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

DEFAULT_CLASSES = [
    "曹新茹",
    "曾巧",
    "陈明远",
    "何嘉欣",
    "李昌润",
    "刘芮娴",
    "宋明达",
    "孙宇天",
    "徐龙桓",
    "张紫健",
]

DEFAULT_NAME_NUMBER_MAP = {
    "徐龙桓": "2025050906",
    "何嘉欣": "2025170101",
    "李昌润": "2025090903",
    "宋明达": "2024050904",
    "陈明远": "2025140901",
    "曹新茹": "2025060902",
    "张紫健": "2025060908",
    "曾巧": "2024170301",
    "刘芮娴": "2024050902",
    "孙宇天": "2025050901",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Classify Word/Excel files to name-based folders by content."
    )
    parser.add_argument(
        "input_dir",
        nargs="?",
        default=".",
        help="Directory containing Word/Excel files (default: current directory).",
    )
    parser.add_argument(
        "--name",
        dest="classes",
        action="append",
        help="Name to match. Repeat to add multiple (default is built-in name list).",
    )
    parser.add_argument(
        "--move",
        action="store_true",
        help="Move files instead of copying.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview classification without creating/changing files.",
    )
    return parser.parse_args()


def canonical_class(name: str) -> str:
    """Return the name as the canonical folder name."""
    return name


def detect_class(file_name: str, text: str, classes: list[str]) -> list[str]:
    """Detect target class by name first, then fall back to number mapping."""
    source = f"{file_name}\n{text}"
    name_hits = [name for name in classes if name in source]
    if name_hits:
        return name_hits

    # If no name is found, use the configured number mapping as a fallback.
    number_hits = [
        name
        for name in classes
        if DEFAULT_NAME_NUMBER_MAP.get(name) and DEFAULT_NAME_NUMBER_MAP[name] in source
    ]
    return number_hits


def read_docx_content(file_path: Path) -> str:
    """Read text content from a .docx file."""
    if Document is None:
        return ""
    try:
        doc = Document(str(file_path))
        return "\n".join([para.text for para in doc.paragraphs])
    except Exception:
        return ""


def read_xlsx_content(file_path: Path) -> str:
    """Read text content from a .xlsx file."""
    if load_workbook is None:
        return ""
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        content = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        content.append(str(cell))
        wb.close()
        return "\n".join(content)
    except Exception:
        return ""


def iter_office_files(root: Path) -> list[Path]:
    """Find all .docx and .xlsx files in the directory."""
    files = []
    for pattern in ["*.docx", "*.xlsx"]:
        files.extend([p for p in root.glob(pattern) if p.is_file()])
    return sorted(files)


def classify_file(md_file: Path, target_dir: Path, move: bool, dry_run: bool) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    target_file = target_dir / md_file.name

    if dry_run:
        action = "MOVE" if move else "COPY"
        print(f"{action}: {md_file} -> {target_file}")
        return

    if move:
        shutil.move(str(md_file), str(target_file))
    else:
        shutil.copy2(md_file, target_file)


def main() -> int:
    args = parse_args()
    root = Path(args.input_dir).expanduser().resolve()

    if not root.exists() or not root.is_dir():
        print(f"Error: directory not found: {root}", file=sys.stderr)
        return 1

    classes_input = args.classes if args.classes else DEFAULT_CLASSES

    if not classes_input:
        print("Error: no valid class names provided.", file=sys.stderr)
        return 1

    office_files = iter_office_files(root)
    if not office_files:
        print("No .docx or .xlsx files found in top-level directory.")
        return 0

    # Check if required libraries are available
    missing_libs = []
    if Document is None:
        missing_libs.append("python-docx")
    if load_workbook is None:
        missing_libs.append("openpyxl")
    if missing_libs:
        print(f"Warning: Missing libraries: {', '.join(missing_libs)}")
        print("Install with: pip install " + " ".join(missing_libs))

    matched = 0
    unmatched = 0
    ambiguous = 0

    for office_file in office_files:
        # Read content based on file type
        if office_file.suffix.lower() == ".docx":
            content = read_docx_content(office_file)
        elif office_file.suffix.lower() == ".xlsx":
            content = read_xlsx_content(office_file)
        else:
            content = ""

        if not content:
            print(f"READ_FAIL: {office_file} (content unavailable, trying filename-only match)", file=sys.stderr)

        hits = detect_class(office_file.name, content, classes_input)

        if len(hits) == 1:
            class_code = canonical_class(hits[0])
            classify_file(office_file, root / class_code, move=args.move, dry_run=args.dry_run)
            matched += 1
            continue

        if len(hits) > 1:
            print(f"AMBIGUOUS: {office_file} -> {', '.join(hits)}", file=sys.stderr)
            ambiguous += 1
            continue

        print(f"UNMATCHED: {office_file}")
        unmatched += 1

    print(
        f"Done. matched={matched}, unmatched={unmatched}, ambiguous={ambiguous}, total={len(office_files)}"
    )
    return 0 if ambiguous == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
